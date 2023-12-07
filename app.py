from flask import Flask, render_template, request, jsonify, redirect, url_for
import pandas as pd
import pickle as pickle
from flask_mysqldb import MySQL
import csv
from openpyxl import load_workbook
import requests

app = Flask(__name__, template_folder='template', static_folder='static')

# MySQL connection
app.config['MYSQL_USER'] = 'root'
app.config['MYSQL_PASSWORD'] = ''
app.config['MYSQL_DB'] = 'todo'
app.config['MYSQL_HOST'] = '127.0.0.1'
app.config['MYSQL_PORT'] = 3306
mysql = MySQL(app)

idcreds = -1

# Routes


@app.route("/login", methods=['GET', 'POST'])
def login():
    global idcreds
    if request.method == 'GET':
        return render_template("login.html"), 200
    if request.method == 'POST':
        json = request.json
        cur = mysql.connection.cursor()
        query = "SELECT * FROM creds WHERE username = %s AND password = %s"
        cur.execute(query, (json["username"], json["password"]))
        data = list(cur.fetchall())
        cur.close()
        if len(data) > 0:
            idcreds = int(data[0][0])
            return "Correct Creds", 200
        else:
            return "Invalid Creds", 500


@app.route("/home", methods=['GET'])
def home():
    cursor = mysql.connection.cursor()
    query = f"SELECT taskid, task_name, date_created, date_completed, completion_status, priority, time_to_complete FROM incomplete WHERE idcreds={idcreds} ORDER BY date_created DESC"
    cursor.execute(query)
    incomplete_task_data = cursor.fetchall()
    incomplete_task_columns = [desc[0] for desc in cursor.description]
    query = f"SELECT taskid, task_name, date_created, date_completed, completion_status, priority, time_to_complete FROM complete WHERE idcreds={idcreds} ORDER BY date_created DESC"
    cursor.execute(query)
    complete_task_data = cursor.fetchall()
    complete_task_columns = [desc[0] for desc in cursor.description]
    cursor.close()
    return render_template("index.html", incomplete_task_data=incomplete_task_data, incomplete_task_columns=incomplete_task_columns, complete_task_data=complete_task_data, complete_task_columns=complete_task_columns)


@app.route("/api/incomplete", methods=['POST', 'GET', 'DELETE', 'PUT'])
def incompleteTasks():
    global idcreds
    if request.method == 'GET':
        cur = mysql.connection.cursor()
        query = "SELECT * FROM incomplete"
        cur.execute(query)
        data = list(cur.fetchall())
        cur.close()
        return data, 200
    if request.method == 'POST':
        json = request.json
        cur = mysql.connection.cursor()
        query = "INSERT INTO incomplete (idcreds, task_name,date_created,date_completed,completion_status,priority,time_to_complete) VALUES (%s, %s, %s, %s, \"Not Completed\", %s, %s)"
        cur.execute(query, (idcreds, json["task_name"], json["date_created"],
                    json["date_completed"], json["priority"], json["time_to_complete"]))
        mysql.connection.commit()
        cur.close()
        return "Incomplete Task Added", 200
    if request.method == 'PUT':
        return render_template("index.html")
    if request.method == 'DELETE':
        json = request.json
        print(json)
        cur = mysql.connection.cursor()
        query = "DELETE FROM incomplete WHERE taskid=%s"
        cur.execute(query, (json["taskid"],))
        mysql.connection.commit()
        cur.close()
        return "Incomplete Task Removed", 200


@app.route("/api/incomplete/clear", methods=['POST'])
def clearIncompleteTasks():
    if request.method == 'POST':
        cur = mysql.connection.cursor()
        query = "DELETE FROM incomplete"
        cur.execute(query)
        mysql.connection.commit()
        cur.close()
        return "Cleared Incomplete Tasks", 200


@app.route("/api/complete", methods=['POST', 'GET', 'DELETE', 'PUT'])
def completeTasks():
    global idcreds
    if request.method == 'GET':
        cur = mysql.connection.cursor()
        query = "SELECT * FROM complete"
        cur.execute(query)
        data = list(cur.fetchall())
        cur.close()
        return data, 200
    if request.method == 'POST':
        json = request.json
        cur = mysql.connection.cursor()
        query = "INSERT INTO complete (taskid, idcreds, task_name,date_created,date_completed,completion_status,priority,time_to_complete) VALUES (%s, %s, %s, %s, %s, \"Completed\", %s, %s)"
        cur.execute(query, (json["taskid"], idcreds, json["task_name"], json["date_created"],
                    json["date_completed"], json["priority"], json["time_to_complete"]))
        mysql.connection.commit()
        cur.close()
        return "Complete Task Added", 200
    if request.method == 'PUT':
        return render_template("index.html")
    if request.method == 'DELETE':
        json = request.json
        cur = mysql.connection.cursor()
        query = "DELETE FROM complete WHERE taskid=%s"
        cur.execute(query, (json["taskid"],))
        mysql.connection.commit()
        cur.close()
        return "Complete Task Removed", 200


@app.route("/api/clear", methods=['POST'])
def clearTasks():
    if request.method == 'POST':
        cur = mysql.connection.cursor()
        query = "DELETE FROM incomplete"
        cur.execute(query)
        query = "DELETE FROM complete"
        cur.execute(query)
        query = "ALTER TABLE incomplete AUTO_INCREMENT = 1"
        cur.execute(query)
        mysql.connection.commit()
        cur.close()
        return "Cleared All Tasks", 200


@app.route("/api/export", methods=['GET'])
def savefile():
    if request.method == 'GET':
        try:
            with open('complete.csv', 'w', newline='') as complete, open('incomplete.csv', 'w', newline='') as incomplete:
                complete_csv_writer = csv.writer(complete)
                incomplete_csv_writer = csv.writer(incomplete)

                # Connect to the database
                cur = mysql.connection.cursor()
                cur.execute('SELECT * FROM complete')
                rows = cur.fetchall()
                complete_csv_writer.writerow(
                    [desc[0] for desc in cur.description])
                complete_csv_writer.writerows(rows)
                cur.execute('SELECT * FROM incomplete')
                rows = cur.fetchall()
                incomplete_csv_writer.writerow(
                    [desc[0] for desc in cur.description])
                incomplete_csv_writer.writerows(rows)

                # Close the database connection
                cur.close()
            return "Saved File into ", 200
        except Exception as e:
            return str(e)


@app.route("/api/upload", methods=['POST'])
def uploadFile():
    file = request.files['file']
    file.save('uploaded_files/' + file.filename)
    workbook = load_workbook('uploaded_files/' + file.filename)
    sheet = workbook.active
    # Process the data as needed
    for row in sheet.iter_rows(min_row=2, values_only=True):
        target_route = 'incompleteTasks'  # Replace with the target route
        target_url = url_for(target_route, _external=True)
        row_data = row[0].split(",")
        data_to_send = {"task_name": row_data[0].strip(),
                        "date_created": row_data[1].strip(),
                        "date_completed": row_data[2].strip(),
                        "completion_status": row_data[3].strip(),
                        "priority": row_data[4].strip(),
                        "time_to_complete": row_data[5].strip()}
        response = requests.post(target_url, json=data_to_send)
        print(f"Row: {row_data}")
    return 'File uploaded and read successfully!', 200


if __name__ == "__main__":
    app.run(debug=True)
